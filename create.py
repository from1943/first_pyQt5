from PyQt5.QtCore import QThread, QMutex, pyqtSignal
import configparser
import openpyxl
from openpyxl.utils import get_column_letter
import copy
import time


# 获取配置
def getConfig():
    cf = configparser.ConfigParser()
    cf.read("static/config.ini")
    titleColumns = cf.get("Default", "titleColumns")
    cityTitle = cf.get("Default", "cityTitle")
    centerTitle = cf.get("Default", "centerTitle")
    return titleColumns, cityTitle, centerTitle


# 表格列标题转序号
def colname_to_num(colname):
    if type(colname) is not str:
        return colname
    col = 0
    power = 1
    for i in range(len(colname)-1, -1, -1):
        ch = colname[i]
        col += (ord(ch)-ord('A')+1)*power
        power *= 26
    return col


# 创建线程锁
create_mut = QMutex()
qmut_2 = QMutex()


# 继承QThread
class Create(QThread):
    def __init__(self, Example):
        super().__init__()
        self.Example = Example

    # 定义信号,生成报表的工作信号执行完成时打开文件窗口选择保存路径
    openSignal = pyqtSignal(object)
    # 定义信号,打印日志
    logSignal = pyqtSignal(str)

    def run(self):
        example = self.Example

        # 生成work_book
        filePath = example.filePath
        sheetname = example.sheetName
        configs = getConfig()
        title = int(configs[0])
        city = configs[1]
        center = configs[2]

        # 处理work_book
        wb_old = openpyxl.load_workbook(filePath, data_only=True)
        wb_new = openpyxl.Workbook()

        print(sheetname)
        sheet = wb_old[sheetname]

        # 先把所有的数据用二维数组储存
        allDatas = []
        datas = list(sheet.iter_rows())[title:]
        for row in datas:
            # 定义行list
            rowDatas = []
            for cell in row:
                rowDatas.append(cell.value)
            if len(rowDatas) > 0:
                allDatas.append(rowDatas)
        # 没有数据就不进行之后的操作
        if len(allDatas) < 1:
            return

        # 根据数据进行分组
        cityNum = colname_to_num(city) - 1
        centerNum = colname_to_num(center) - 1
        # 根据目标列排序
        allDatas.sort(key=lambda k: ((k[cityNum] is None, k[cityNum] == "", k[cityNum]), (k[centerNum] is None, k[centerNum] == "", k[centerNum])))
        # 将数据进行分组
        newSheetDict = {}
        for row in allDatas:
            # 用城市加上中心作为sheet名称，不能重复
            currentSheetName = str(row[cityNum]) + str(row[centerNum])
            if currentSheetName not in newSheetDict:
                newSheetDict[currentSheetName] = []
            # 将数据放进对应的key下面
            newSheetDict[currentSheetName].append(row)

        if len(newSheetDict) < 1:
            return

        # 遍历已经进行了分组的dict
        for key, value in newSheetDict.items():
            sheet_new = wb_new.create_sheet(key)
            # 以下是复制标题逻辑（包含基本样式）
            # tab颜色
            sheet_new.sheet_properties.tabColor = sheet.sheet_properties.tabColor
            # 开始处理合并单元格形式为“(<CellRange A1：A4>,)，替换掉(<CellRange 和 >,)' 找到合并单元格
            wm = list(sheet.merged_cells)
            if len(wm) > 0:
                for i in range(0, len(wm)):
                    cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                    sheet_new.merge_cells(cell2)

            # 复制标题
            items = list(sheet.iter_rows())[0:title]
            for i, row in enumerate(items):
                print(str(i))
                print(str(row))
                # 发送信号打印日志
                self.logSignal.emit(str(row))
                sheet_new.row_dimensions[i + 1].height = sheet.row_dimensions[i + 1].height
                for j, cell in enumerate(row):
                    sheet_new.column_dimensions[get_column_letter(j + 1)].width = sheet.column_dimensions[
                        get_column_letter(j + 1)].width
                    sheet_new.cell(row=i + 1, column=j + 1, value=cell.value)

                    # 设置单元格格式
                    source_cell = sheet.cell(i + 1, j + 1)
                    target_cell = sheet_new.cell(i + 1, j + 1)
                    target_cell.fill = copy.copy(source_cell.fill)
                    if source_cell.has_style:
                        target_cell._style = copy.copy(source_cell._style)
                        target_cell.font = copy.copy(source_cell.font)
                        target_cell.border = copy.copy(source_cell.border)
                        target_cell.fill = copy.copy(source_cell.fill)
                        target_cell.number_format = copy.copy(source_cell.number_format)
                        target_cell.protection = copy.copy(source_cell.protection)
                        target_cell.alignment = copy.copy(source_cell.alignment)

            # 以下是处理数据从标题行之后的行开始填充
            for i, row in enumerate(value):
                for j, cell in enumerate(row):
                    if cell is not None:
                        sheet_new.cell(row=i + 1 + title, column=j + 1, value=cell)

            if 'Sheet' in wb_new.sheetnames:
                del wb_new['Sheet']
        # 发送信号打印日志 success
        self.logSignal.emit('success!')

        # 发送结束生成任务
        self.openSignal.emit(wb_new)

        # wb_new.save(file_path[0])
        # wb_new.close()
        wb_old.close()


class Progress(QThread):
    def __init__(self, Example):
        super().__init__()
        self.Example = Example

    def run(self):
        example = self.Example
        while example.isActive:
            time.sleep(1)
            setp = example.ui.progressBar.value()
            if (setp + 5) <= 90:
                nextStep = setp + 5
                example.ui.progressBar.setValue(nextStep)

