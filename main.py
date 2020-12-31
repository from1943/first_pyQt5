import sys
import time

from PyQt5.QtGui import QPixmap
from PyQt5.QtWidgets import *
import os
import export
import configparser
import openpyxl

from create import Create, Progress


# 获取配置
def get_config():
    cf = configparser.ConfigParser()
    cf.read("static/config.ini")
    titleColumns = cf.get("Default", "titleColumns")
    cityTitle = cf.get("Default", "cityTitle")
    centerTitle = cf.get("Default", "centerTitle")
    return titleColumns, cityTitle, centerTitle


# 比对配置文件和用户填写的值是否一致,如果不一致就修改配置文件
def check_ini(title, city, center):
    configs = get_config()
    if not (title == configs[0] and city.upper() == configs[1] and center.upper() == configs[2]):
        cf = configparser.ConfigParser()
        cf.read("static/config.ini")
        cf.set("Default", "titleColumns", title)
        cf.set("Default", "cityTitle", city.upper())
        cf.set("Default", "centerTitle", center.upper())
        cf.write(open("static/config.ini", "r+"))


class Example(QMainWindow):
    def __init__(self):
        self.app = QApplication(sys.argv)
        super().__init__()
        self.ui = export.Ui_MainWindow()
        self.ui.setupUi(self)
        # 获取当前程序的位置，选择文件时使用
        self.cwd = os.getcwd()
        # 存一个当前页面，翻页和返回时可以使用
        self.currentIndex = 0
        # 储存文件完整路径
        self.filePath = None
        # 储存sheetName
        self.sheetName = None
        # 储存任务执行状态
        self.isActive = True

        # 设置工作线程和主线程绑定
        self.progress_thread = None
        self.workbook_thread = None

        # 初始化
        self.init_ui()

    # confirm事件--第一页翻页
    def click_confirm(self):
        # 第一页翻页时需要检查是否选择了文件，还需要将文件中的sheet读取出来
        if self.ui.label_fileName.text() is None or self.ui.label_fileName.text() == '' or self.ui.label_fileName.text() == '请先选择文件':
            self.ui.label_fileName.setText("请先选择文件")
            return
        # 将选中的文件所有sheet读取出来，放进下拉框
        filePath = self.filePath
        wb = openpyxl.load_workbook(filePath)
        sheetnames = wb.sheetnames
        self.ui.comboBox.addItems(sheetnames)
        # 把第一个sheetName赋值给lable
        self.ui.label_5.setText(sheetnames[0])
        # 翻页
        targetIndex = self.currentIndex + 1
        self.currentIndex = targetIndex
        self.ui.stackedWidget.setCurrentIndex(targetIndex)

    # confirm事件--第二页翻页
    def click_confirm_2(self):
        # 第二页翻页时需要下拉框label是否正确显示
        labelValue = self.ui.label_5.text()
        if labelValue is None or labelValue == '' or labelValue == '请选择sheet':
            self.ui.label_5.setText("请选择sheet")
        self.sheetName = labelValue
        # 翻页
        targetIndex = self.currentIndex + 1
        self.currentIndex = targetIndex
        self.ui.stackedWidget.setCurrentIndex(targetIndex)

    # confirm事件--第三页翻页
    def click_confirm_3(self):
        # 第三页翻页时需要检查配置是否正确
        titleColumns = self.ui.line_title.text()
        if titleColumns is None or titleColumns == '' or titleColumns == '必填项！':
            self.ui.line_title.setText('必填项！')
            return
        elif not titleColumns.isdigit():
            self.ui.line_title.setText('必须为整数！')
            return
        cityTitle = self.ui.line_city.text()
        if cityTitle is None or cityTitle == '' or cityTitle == '必填项！':
            self.ui.line_city.setText('必填项！')
            return
        elif not cityTitle.isalpha():
            self.ui.line_city.setText('必须为字母！')
            return
        centerTitle = self.ui.line_center.text()
        if centerTitle is None or centerTitle == '' or centerTitle == '必填项！':
            self.ui.line_center.setText('必填项！')
            return
        elif not centerTitle.isalpha():
            self.ui.line_center.setText('必须为字母！')
            return
        # 如果与配置文件不同则需要修改配置文件
        check_ini(titleColumns, cityTitle, centerTitle)
        targetIndex = self.currentIndex + 1
        self.currentIndex = targetIndex
        # 第四页的配置展示填充
        configs = get_config()
        self.ui.label_title_3.setText(configs[0])
        self.ui.label_city_3.setText(configs[1])
        self.ui.label_center_3.setText(configs[2])
        # 翻页
        self.ui.stackedWidget.setCurrentIndex(targetIndex)

    # 第四页确认
    def click_confirm_4(self):
        targetIndex = self.currentIndex + 1
        self.currentIndex = targetIndex
        # 翻页
        self.ui.stackedWidget.setCurrentIndex(targetIndex)
        # 将任务状态重置为True
        self.isActive = True
        self.ui.progressBar.setValue(5)
        # 调用进度条任务
        self.progress_thread = Progress(self)
        self.progress_thread.start()
        # 调用生成任务
        self.workbook_thread = Create(self)
        # 绑定信号槽
        self.workbook_thread.openSignal.connect(self.choose_file_path)
        self.workbook_thread.logSignal.connect(self.text_log_run)
        self.workbook_thread.start()

    def choose_file_path(self, wb_new):
        self.isActive = False
        # 将进度条设置为100
        self.ui.progressBar.setValue(100)
        # 将Back和Finish设置为可用
        self.ui.button_back_4.setEnabled(True)
        self.ui.button_confirm_5.setEnabled(True)
        file_path = QFileDialog.getSaveFileName(self, "保存文件", self.cwd + "/未命名", "xlsx files (*.xlsx)")
        print(file_path)
        wb_new.save(file_path[0])
        wb_new.close()

    # 关闭UI
    def click_close(self):
        self.close()

    # back事件--返回上一页
    def click_back(self):
        if self.currentIndex == 0:
            return
        targetIndex = self.currentIndex - 1
        self.currentIndex = targetIndex
        self.ui.stackedWidget.setCurrentIndex(targetIndex)

    # 菜单点击(打开文件资源管理器)
    def menu_click_success(self):
        file_name = QFileDialog.getOpenFileName(self, "选择excel文件", self.cwd, "Xlsx files(*.xlsx);;Xls files(*.xls)")
        print(file_name[0])
        if file_name[0] != '':
            # 保存文件绝对路径
            self.filePath = file_name[0]
            self.ui.label_fileName_3.setText(file_name[0])
            # 加载文件名称
            name = file_name[0].split('/')[-1]
            prix = file_name[0].split('.')[-1]
            self.ui.label_fileName.setText(name)
            # 加载图标 根据文件后缀分辨是xlsx还是xls
            icon = QPixmap('static/xlsx.png') if prix == 'xlsx' else QPixmap('static/xls.png')
            self.ui.label_img.setPixmap(icon)

    # 下拉框选中信号，改变label显示内容
    def select_label(self, value):
        self.ui.label_5.setText(value)
        time.sleep(0.2)

    # 日志打印
    def text_log_run(self, content):
        # 打印日志
        self.ui.textEdit.append(content)

    # ui初始化
    def init_ui(self):
        # 设置Open菜单选择事件
        self.ui.actionOpen.triggered.connect(self.menu_click_success)
        # 设置Open菜单的快捷键
        self.ui.actionOpen.setShortcut('Ctrl+O')
        # 设置confirm
        self.ui.button_confirm.clicked.connect(self.click_confirm)
        self.ui.button_confirm_4.clicked.connect(self.click_confirm_2)
        self.ui.button_confirm_2.clicked.connect(self.click_confirm_3)
        self.ui.button_confirm_3.clicked.connect(self.click_confirm_4)
        self.ui.button_confirm_5.clicked.connect(self.click_close)
        # 设置back
        self.ui.button_back.clicked.connect(self.click_back)
        self.ui.button_back_3.clicked.connect(self.click_back)
        self.ui.button_back_2.clicked.connect(self.click_back)
        self.ui.button_back_4.clicked.connect(self.click_back)
        # 读取配置文件——获取默认标题行内容行设置
        configs = get_config()
        self.ui.line_title.setText(configs[0])
        self.ui.line_city.setText(configs[1])
        self.ui.line_center.setText(configs[2])
        # 下拉框槽函数
        self.ui.comboBox.activated[str].connect(self.select_label)
        self.show()


if __name__ == '__main__':
    e = Example()
    sys.exit(e.app.exec())
